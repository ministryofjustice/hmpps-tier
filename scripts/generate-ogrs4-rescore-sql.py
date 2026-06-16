#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.14,<4"
# dependencies = [
#     "openpyxl>=3.1,<4",
# ]
# ///
import argparse
import sys
from collections.abc import Callable, Iterable, Sequence
from datetime import date, datetime, timezone
from decimal import Decimal
from itertools import batched
from openpyxl import load_workbook
from pathlib import Path
from zoneinfo import ZoneInfo

TABLE = "public.ogrs4_rescored_assessment"
LONDON = ZoneInfo("Europe/London")
DEFAULT_BATCH_SIZE = 1_000

TARGET_COLUMNS = [
    "crn",
    "completed_date",
    "arp_score",
    "arp_is_dynamic",
    "arp_band",
    "csrp_score",
    "csrp_is_dynamic",
    "csrp_band",
    "dc_srp_score",
    "dc_srp_band",
    "iic_srp_score",
    "iic_srp_band",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert an OGRS4 rescored assessment XLSX file into a batched Postgres INSERT SQL script."
    )
    parser.add_argument("xlsx", type=Path, help="Input XLSX file")
    parser.add_argument("sql", type=Path, help="Output SQL file")
    parser.add_argument("--batch-size", type=int, default=DEFAULT_BATCH_SIZE)
    return parser.parse_args()


def log(message: str) -> None:
    now = datetime.now(timezone.utc).isoformat()
    print(f"[{now}] {message}", flush=True)


def is_blank(value: object) -> bool:
    return value is None or str(value).strip() == ""


def parse_text(value: object) -> str | None:
    return None if is_blank(value) else str(value).strip()


def parse_number(value: object) -> Decimal | None:
    return None if is_blank(value) else Decimal(str(value))


def parse_dynamic(value: object) -> bool:
    return str(value).strip().upper() == "DYNAMIC"


def parse_band(value: object) -> str | None:
    if is_blank(value):
        return None
    if value not in ("Low", "Medium", "High", "Very High", "Not Applicable"):
        raise ValueError(f"{value} is not valid")
    return value


def parse_completed_date(value: object) -> datetime:
    if isinstance(value, datetime):
        return value.replace(tzinfo=LONDON)

    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time(), tzinfo=LONDON)

    cleaned = " ".join(str(value).split())
    return datetime.strptime(cleaned, "%d/%m/%Y %H:%M:%S").replace(tzinfo=LONDON)


SOURCE_COLUMNS: list[tuple[str, Callable[[object], object]]] = [
    ("CMS_PROB_NUMBER", parse_text),
    ("DATE_COMPLETED", parse_completed_date),
    ("ARP", parse_number),
    ("ARP_STATIC_DYNAMIC", parse_dynamic),
    ("ARP_BAND", parse_band),
    ("CSRP", parse_number),
    ("CSRP_STATIC_DYNAMIC", parse_dynamic),
    ("CSRP_BAND", parse_band),
    ("DC_SRP_PERCENTAGE", parse_number),
    ("DC_SRP_BAND", parse_band),
    ("IIC_SRP_SCORE", parse_number),
    ("IIC_SRP_BAND", parse_band),
]


def sql_literal(value: object) -> str:
    if value is None:
        return "NULL"

    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"

    if isinstance(value, Decimal):
        return str(value)

    if isinstance(value, datetime):
        return f"'{value.isoformat(sep=' ', timespec='seconds')}'"

    if isinstance(value, date):
        return f"'{value.isoformat()}'"

    escaped = str(value).replace("'", "''")
    return f"'{escaped}'"


def read_assessments(xlsx_path: Path) -> Iterable[list[object]]:
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)

    headers = {
        str(value).strip(): index
        for index, value in enumerate(next(rows))
        if value is not None
    }

    missing = [name for name, _ in SOURCE_COLUMNS if name not in headers]
    if missing:
        raise ValueError(f"Missing expected columns: {', '.join(missing)}")

    for row in rows:
        if all(is_blank(value) for value in row):
            continue

        yield [
            parser(row[headers[source_column]])
            for source_column, parser in SOURCE_COLUMNS
        ]


def write_insert(sql, rows: Sequence[Sequence[object]]) -> None:
    sql.write("BEGIN;\n")
    sql.write(f"INSERT INTO {TABLE} (\n")
    sql.write("    " + ",\n    ".join(TARGET_COLUMNS))
    sql.write("\n)\nVALUES\n")

    sql.write(",\n".join(
        "    (" + ", ".join(sql_literal(value) for value in row) + ")"
        for row in rows
    ))

    sql.write(";\n")
    sql.write("COMMIT;\n\n")


def write_sql(assessments: Iterable[list[object]], sql_path: Path, batch_size: int) -> int:
    row_count = 0

    with sql_path.open("w", encoding="utf-8") as sql:
        for batch in batched(assessments, batch_size):
            write_insert(sql, batch)
            row_count += len(batch)

        if row_count == 0:
            sql.write("-- No rows found.\n")

    return row_count


def main() -> int:
    args = parse_args()

    if args.batch_size <= 0:
        print("batch-size must be > 0", file=sys.stderr)
        return 2

    if not args.xlsx.exists():
        print(f"Input XLSX file does not exist: {args.xlsx}", file=sys.stderr)
        return 2

    log(f"Reading {args.xlsx}")
    log(f"Writing SQL to {args.sql} with batch_size={args.batch_size}")

    row_count = write_sql(
        assessments=read_assessments(args.xlsx),
        sql_path=args.sql,
        batch_size=args.batch_size,
    )

    log(f"Finished. Wrote {row_count} rows.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
